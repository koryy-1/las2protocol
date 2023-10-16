import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import coordinate_to_tuple

def exceeds_threshold(calculation_table):
    max_percent_diff = max(
        abs(calculation_table[5][2]), abs(calculation_table[5][3]), abs(calculation_table[5][4]),
        abs(calculation_table[6][2]), abs(calculation_table[6][3]), abs(calculation_table[6][4])
    )
    return max_percent_diff >= 5
    
def find_temperature_drop_point(temp_data, tolerance_degrees):
    T_max_index = np.argmax(temp_data)
    T_max = temp_data[T_max_index]
    T_max_index_last = T_max_index
    for i in range(T_max_index, len(temp_data)):
        # Вычисляем разницу между максимальным и текущим значением
        temp_difference = T_max - temp_data[i]
        if T_max == temp_data[i]:
            T_max_index_last = i
        
        # Если разница превышает заданный порог (tolerance_degrees), это может быть началом уменьшения
        if temp_difference >= tolerance_degrees:
            return T_max_index_last, temp_data[T_max_index_last]  # Возвращаем индекс и значение температуры
    return None, None  # Если уменьшения не обнаружено

def find_temperature_rise_point(temp_data, tolerance_degrees):
    for i in range(0, len(temp_data) - 1):
        # Вычисляем разницу между следующим и текущим значением
        temp_difference = temp_data[i + 1] - temp_data[i]
        
        # Если разница превышает заданный порог (tolerance_degrees), это может быть началом увеличения
        if temp_difference >= tolerance_degrees:
            return i, temp_data[i]  # Возвращаем индекс и значение температуры
    print('program not defined start of heating')
    return None, None  # Если увеличения не обнаружено

def find_temperature_rise_point_right(temp_data, tolerance_degrees):
    for i in range(len(temp_data) - 1, 0, -1):
        # Вычисляем разницу между текущим и следующим значением
        temp_difference = temp_data[i - 1] - temp_data[i]

        # Если разница превышает заданный порог (tolerance_degrees), это может быть началом увеличения
        if temp_difference >= tolerance_degrees:
            return i, temp_data[i]  # Возвращаем индекс и значение температуры
    print('program not defined start of heating in right side of temp plot')
    return None, None  # Если увеличения не обнаружено справа


def find_min_variance_interval(data, interval_length=60):
    # min_variance = float('inf')  # Начальное значение минимальной дисперсии
    # min_variance_interval = []  # Инициализация интервала с минимальной дисперсией
    # min_variance_index = 0  # Индекс начала интервала с минимальной дисперсией
    variance_arr = []

    for i in range(0, len(data) - interval_length + 1, interval_length):
        interval = data[i:i + interval_length]  # Выбираем интервал длиной 60 отсчетов
        variance = np.var(interval)  # Вычисляем дисперсию интервала
        variance_arr.append(variance)

        # # Если текущая дисперсия меньше минимальной, обновляем значения
        # if variance < min_variance:
        #     min_variance = variance
        #     min_variance_interval = interval
        #     min_variance_index = i

    return variance_arr

def insert_formula_into_cell(sheet, cell, formula):
    sheet[cell].formula = formula

def smoothing_function(data, window_size, count):
    # via excel (crutch)
    # workbook = openpyxl.Workbook()
    # sheet = workbook.active

    # formula = '=IF(ROW()<2,NA(),AVERAGE(OFFSET(C2,-120,0,240,1)))'

    # # Найдем ячейку, в которую нужно вставить формулу
    # formula_cell = 'A1'  # Замените на актуальное значение
    # sheet[formula_cell].value = formula

    # # Получаем координаты ячейки с формулой
    # formula_row, formula_col = coordinate_to_tuple(formula_cell)

    # output_array = []

    # for i in range(len(data)):
    #     # Устанавливаем значение переменной C2 равным текущему элементу массива
    #     sheet.cell(row=formula_row, column=formula_col, value=data[i])

    #     # Вычисляем формулу
    #     result = sheet.cell(row=formula_row, column=formula_col).value

    #     # Добавляем результат в выходной массив
    #     output_array.append(result)

    # workbook.close()

    # return np.array(output_array)
    

    ### median smooth
    smoothed_data = data
    for _ in range(count):
        smoothed_data = np.convolve(smoothed_data, np.ones(window_size)/window_size, mode='valid')
    return smoothed_data


    ### default
    # smoothed_data = data
    # for _ in range(count):
    #     smoothed_data = pd.Series(smoothed_data).rolling(window=window_size).mean().iloc[window_size-1:].values
    # return smoothed_data
